from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
import json
from typing import Any
from uuid import UUID
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ModuleNotFoundError:
    Workbook = None
    Font = None
    PatternFill = None
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from ..database import DbSession
from ..models.audit_log import AuditLog
from ..models.documento import Documento
from ..models.upload import Upload


@dataclass(frozen=True, slots=True)
class ExportColumn:
    key: str
    header: str
    width: int = 20


@dataclass(frozen=True, slots=True)
class ExportSheet:
    title: str
    columns: tuple[ExportColumn, ...]
    rows: list[dict[str, Any]]
    severity_key: str | None = None


@dataclass(frozen=True, slots=True)
class GeneratedExport:
    content: bytes
    total_registros: int


DOCUMENT_EXPORT_COLUMNS = (
    ExportColumn("upload_id", "Upload ID", 38),
    ExportColumn("documento_id", "Documento ID", 38),
    ExportColumn("nome_arquivo", "Arquivo", 26),
    ExportColumn("numero_nf", "Numero NF", 18),
    ExportColumn("cnpj_emitente", "CNPJ Emitente", 20),
    ExportColumn("cnpj_destinatario", "CNPJ Destinatario", 20),
    ExportColumn("data_emissao", "Data Emissao", 14),
    ExportColumn("data_pagamento", "Data Pagamento", 16),
    ExportColumn("valor_total", "Valor Total", 14),
    ExportColumn("aprovador", "Aprovador", 24),
    ExportColumn("descricao", "Descricao", 40),
    ExportColumn("status", "Status", 16),
    ExportColumn("resumo", "Resumo", 44),
    ExportColumn("flag_codigo", "Flag Codigo", 18),
    ExportColumn("flag_descricao", "Flag Descricao", 42),
    ExportColumn("flag_severidade", "Flag Severidade", 16),
    ExportColumn("flag_detectada_em", "Flag Detectada Em", 24),
)

AUDIT_LOG_EXPORT_COLUMNS = (
    ExportColumn("id", "ID", 38),
    ExportColumn("evento", "Evento", 28),
    ExportColumn("entidade_tipo", "Entidade Tipo", 20),
    ExportColumn("entidade_id", "Entidade ID", 38),
    ExportColumn("usuario", "Usuario", 24),
    ExportColumn("ip", "IP", 18),
    ExportColumn("payload", "Payload", 60),
    ExportColumn("timestamp", "Timestamp", 24),
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937") if PatternFill else None
SEVERITY_FILLS = (
    {
        "CRITICA": PatternFill(fill_type="solid", fgColor="FECACA"),
        "ALTA": PatternFill(fill_type="solid", fgColor="FED7AA"),
        "MEDIA": PatternFill(fill_type="solid", fgColor="FEF3C7"),
    }
    if PatternFill
    else {}
)


def _build_summary(status: str, anomaly_count: int) -> str:
    normalized_status = status.casefold()
    if normalized_status == "erro":
        return "Processamento interrompido com erro."
    if normalized_status in {"pendente", "processando"}:
        return "Upload recebido e aguardando pipeline de processamento."
    if anomaly_count:
        return f"{anomaly_count} anomalia(s) detectada(s) para revisao."
    return "Processamento concluido sem anomalias."


def _select_latest_document(upload: Upload) -> Documento | None:
    if not upload.documentos:
        return None
    return max(upload.documentos, key=lambda documento: documento.criado_em or datetime.min)


def _serialize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=_serialize_value)
    return str(value)


def _build_document_export_rows(
    db: DbSession, *, somente_com_anomalias: bool = False
) -> list[dict[str, Any]]:
    uploads = db.scalars(
        select(Upload)
        .options(selectinload(Upload.documentos).selectinload(Documento.anomalias))
        .order_by(Upload.criado_em.desc())
    ).all()

    rows: list[dict[str, Any]] = []
    for upload in uploads:
        documento = _select_latest_document(upload)
        if documento is None:
            if somente_com_anomalias:
                continue
            rows.append(
                {
                    "upload_id": upload.id,
                    "documento_id": None,
                    "nome_arquivo": upload.nome_arquivo,
                    "numero_nf": None,
                    "cnpj_emitente": None,
                    "cnpj_destinatario": None,
                    "data_emissao": None,
                    "data_pagamento": None,
                    "valor_total": None,
                    "aprovador": None,
                    "descricao": None,
                    "status": upload.status,
                    "resumo": _build_summary(upload.status, 0),
                    "flag_codigo": None,
                    "flag_descricao": None,
                    "flag_severidade": None,
                    "flag_detectada_em": None,
                }
            )
            continue

        anomalias = sorted(documento.anomalias, key=lambda item: item.criado_em or datetime.min)
        if somente_com_anomalias and not anomalias:
            continue

        base_row = {
            "upload_id": upload.id,
            "documento_id": documento.id,
            "nome_arquivo": upload.nome_arquivo,
            "numero_nf": documento.numero_nf,
            "cnpj_emitente": documento.cnpj_emitente,
            "cnpj_destinatario": documento.cnpj_destinatario,
            "data_emissao": documento.data_emissao,
            "data_pagamento": documento.data_pagamento,
            "valor_total": documento.valor_total,
            "aprovador": documento.aprovador,
            "descricao": documento.descricao,
            "status": documento.status_extracao,
            "resumo": _build_summary(documento.status_extracao, len(anomalias)),
        }

        if not anomalias:
            rows.append(
                {
                    **base_row,
                    "flag_codigo": None,
                    "flag_descricao": None,
                    "flag_severidade": None,
                    "flag_detectada_em": None,
                }
            )
            continue

        for anomalia in anomalias:
            rows.append(
                {
                    **base_row,
                    "flag_codigo": anomalia.codigo,
                    "flag_descricao": anomalia.descricao,
                    "flag_severidade": anomalia.severidade,
                    "flag_detectada_em": anomalia.criado_em,
                }
            )

    return rows


def _build_audit_log_export_rows(db: DbSession) -> list[dict[str, Any]]:
    audit_logs = db.scalars(select(AuditLog).order_by(AuditLog.timestamp.desc())).all()
    return [
        {
            "id": audit_log.id,
            "evento": audit_log.evento,
            "entidade_tipo": audit_log.entidade_tipo,
            "entidade_id": audit_log.entidade_id,
            "usuario": audit_log.usuario,
            "ip": audit_log.ip,
            "payload": audit_log.payload,
            "timestamp": audit_log.timestamp,
        }
        for audit_log in audit_logs
    ]


def generate_csv_bytes(rows: list[dict[str, Any]], columns: tuple[ExportColumn, ...]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, delimiter=";", lineterminator="\r\n")
    writer.writerow([column.header for column in columns])

    for row in rows:
        writer.writerow([_serialize_value(row.get(column.key)) for column in columns])

    return output.getvalue().encode("utf-8-sig")


def generate_excel_bytes(sheets: tuple[ExportSheet, ...]) -> bytes:
    if Workbook is None or Font is None or HEADER_FILL is None:
        return _generate_excel_bytes_fallback(sheets)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet in sheets:
        worksheet = workbook.create_sheet(sheet.title)
        worksheet.append([column.header for column in sheet.columns])
        worksheet.freeze_panes = "A2"

        for header_cell in worksheet[1]:
            header_cell.font = Font(bold=True, color="FFFFFF")
            header_cell.fill = HEADER_FILL

        key_index = {column.key: index + 1 for index, column in enumerate(sheet.columns)}
        severity_column_index = key_index.get(sheet.severity_key) if sheet.severity_key else None

        for row in sheet.rows:
            worksheet.append([_serialize_value(row.get(column.key)) for column in sheet.columns])

            if severity_column_index is None:
                continue

            severity_cell = worksheet.cell(row=worksheet.max_row, column=severity_column_index)
            severity_value = str(severity_cell.value or "").upper()
            fill = SEVERITY_FILLS.get(severity_value)
            if fill is not None:
                severity_cell.fill = fill
                severity_cell.font = Font(bold=True)

        worksheet.auto_filter.ref = worksheet.dimensions
        for index, column in enumerate(sheet.columns, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = column.width

    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _generate_excel_bytes_fallback(sheets: tuple[ExportSheet, ...]) -> bytes:
    content = BytesIO()
    created_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    with ZipFile(content, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", _build_content_types_xml(len(sheets)))
        archive.writestr("_rels/.rels", _build_root_relationships_xml())
        archive.writestr("docProps/app.xml", _build_app_properties_xml(sheets))
        archive.writestr("docProps/core.xml", _build_core_properties_xml(created_at))
        archive.writestr("xl/workbook.xml", _build_workbook_xml(sheets))
        archive.writestr("xl/_rels/workbook.xml.rels", _build_workbook_relationships_xml(len(sheets)))
        archive.writestr("xl/styles.xml", _build_styles_xml())

        for index, sheet in enumerate(sheets, start=1):
            archive.writestr(f"xl/worksheets/sheet{index}.xml", _build_worksheet_xml(sheet))

    return content.getvalue()


def _build_content_types_xml(sheet_count: int) -> str:
    worksheet_overrides = "".join(
        (
            f'<Override PartName="/xl/worksheets/sheet{index}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
        for index in range(1, sheet_count + 1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        f"{worksheet_overrides}"
        '<Override PartName="/xl/styles.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/docProps/core.xml" '
        'ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        "</Types>"
    )


def _build_root_relationships_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" '
        'Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" '
        'Target="docProps/app.xml"/>'
        "</Relationships>"
    )


def _build_app_properties_xml(sheets: tuple[ExportSheet, ...]) -> str:
    titles = "".join(f"<vt:lpstr>{escape(sheet.title)}</vt:lpstr>" for sheet in sheets)
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        "<Application>DocAudit</Application>"
        "<HeadingPairs>"
        '<vt:vector size="2" baseType="variant">'
        "<vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant>"
        f"<vt:variant><vt:i4>{len(sheets)}</vt:i4></vt:variant>"
        "</vt:vector>"
        "</HeadingPairs>"
        "<TitlesOfParts>"
        f'<vt:vector size="{len(sheets)}" baseType="lpstr">{titles}</vt:vector>'
        "</TitlesOfParts>"
        "</Properties>"
    )


def _build_core_properties_xml(created_at: str) -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        "<dc:creator>DocAudit</dc:creator>"
        "<cp:lastModifiedBy>DocAudit</cp:lastModifiedBy>"
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{created_at}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{created_at}</dcterms:modified>'
        "</cp:coreProperties>"
    )


def _build_workbook_xml(sheets: tuple[ExportSheet, ...]) -> str:
    sheet_nodes = "".join(
        (
            f'<sheet name="{escape(sheet.title)}" sheetId="{index}" r:id="rId{index}"/>'
            for index, sheet in enumerate(sheets, start=1)
        )
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{sheet_nodes}</sheets>"
        "</workbook>"
    )


def _build_workbook_relationships_xml(sheet_count: int) -> str:
    sheet_relationships = "".join(
        (
            f'<Relationship Id="rId{index}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{index}.xml"/>'
        )
        for index in range(1, sheet_count + 1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f"{sheet_relationships}"
        f'<Relationship Id="rId{sheet_count + 1}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        'Target="styles.xml"/>'
        "</Relationships>"
    )


def _build_styles_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="3">'
        '<font><sz val="11"/><name val="Calibri"/><family val="2"/></font>'
        '<font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/><family val="2"/></font>'
        '<font><b/><sz val="11"/><name val="Calibri"/><family val="2"/></font>'
        "</fonts>"
        '<fills count="6">'
        '<fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FF1F2937"/><bgColor indexed="64"/></patternFill></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FFFECACA"/><bgColor indexed="64"/></patternFill></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FFFED7AA"/><bgColor indexed="64"/></patternFill></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FFFEF3C7"/><bgColor indexed="64"/></patternFill></fill>'
        "</fills>"
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="5">'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        '<xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1"/>'
        '<xf numFmtId="0" fontId="2" fillId="3" borderId="0" xfId="0" applyFont="1" applyFill="1"/>'
        '<xf numFmtId="0" fontId="2" fillId="4" borderId="0" xfId="0" applyFont="1" applyFill="1"/>'
        '<xf numFmtId="0" fontId="2" fillId="5" borderId="0" xfId="0" applyFont="1" applyFill="1"/>'
        "</cellXfs>"
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        "</styleSheet>"
    )


def _build_worksheet_xml(sheet: ExportSheet) -> str:
    row_count = len(sheet.rows) + 1
    column_count = len(sheet.columns)
    max_ref = f"{_xlsx_column_letter(column_count)}{row_count}"
    header_cells = "".join(
        _build_inline_string_cell(
            row_index=1,
            column_index=index,
            value=column.header,
            style_id=1,
        )
        for index, column in enumerate(sheet.columns, start=1)
    )
    data_rows = []
    severity_column_index = next(
        (index for index, column in enumerate(sheet.columns, start=1) if column.key == sheet.severity_key),
        None,
    )

    for row_index, row in enumerate(sheet.rows, start=2):
        row_cells = []
        for column_index, column in enumerate(sheet.columns, start=1):
            style_id = 0
            if severity_column_index == column_index:
                severity_value = str(row.get(column.key) or "").upper()
                style_id = {
                    "CRITICA": 2,
                    "ALTA": 3,
                    "MEDIA": 4,
                }.get(severity_value, 0)
            row_cells.append(
                _build_inline_string_cell(
                    row_index=row_index,
                    column_index=column_index,
                    value=_serialize_value(row.get(column.key)),
                    style_id=style_id,
                )
            )
        data_rows.append(f'<row r="{row_index}">{"".join(row_cells)}</row>')

    cols = "".join(
        (
            f'<col min="{index}" max="{index}" width="{column.width}" customWidth="1"/>'
            for index, column in enumerate(sheet.columns, start=1)
        )
    )
    auto_filter = f'<autoFilter ref="A1:{max_ref}"/>'
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="A1:{max_ref}"/>'
        '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" '
        'activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>'
        f"<cols>{cols}</cols>"
        "<sheetData>"
        f'<row r="1">{header_cells}</row>'
        f'{"".join(data_rows)}'
        "</sheetData>"
        f"{auto_filter}"
        "</worksheet>"
    )


def _build_inline_string_cell(row_index: int, column_index: int, value: str, style_id: int = 0) -> str:
    sanitized_value = escape(value)
    style_attr = f' s="{style_id}"' if style_id else ""
    return (
        f'<c r="{_xlsx_column_letter(column_index)}{row_index}" t="inlineStr"{style_attr}>'
        f"<is><t>{sanitized_value}</t></is>"
        "</c>"
    )


def _xlsx_column_letter(index: int) -> str:
    letters: list[str] = []
    remaining = index
    while remaining:
        remaining, remainder = divmod(remaining - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def export_documentos_csv(
    db: DbSession, *, somente_com_anomalias: bool = False
) -> GeneratedExport:
    rows = _build_document_export_rows(db, somente_com_anomalias=somente_com_anomalias)
    return GeneratedExport(
        content=generate_csv_bytes(rows, DOCUMENT_EXPORT_COLUMNS),
        total_registros=len(rows),
    )


def export_documentos_excel(
    db: DbSession, *, somente_com_anomalias: bool = False
) -> GeneratedExport:
    document_rows = _build_document_export_rows(db, somente_com_anomalias=somente_com_anomalias)
    audit_log_rows = _build_audit_log_export_rows(db)
    return GeneratedExport(
        content=generate_excel_bytes(
            (
                ExportSheet(
                    title="Documentos",
                    columns=DOCUMENT_EXPORT_COLUMNS,
                    rows=document_rows,
                    severity_key="flag_severidade",
                ),
                ExportSheet(
                    title="Log Auditoria",
                    columns=AUDIT_LOG_EXPORT_COLUMNS,
                    rows=audit_log_rows,
                ),
            )
        ),
        total_registros=len(document_rows),
    )


def export_audit_log_csv(db: DbSession) -> GeneratedExport:
    rows = _build_audit_log_export_rows(db)
    return GeneratedExport(
        content=generate_csv_bytes(rows, AUDIT_LOG_EXPORT_COLUMNS),
        total_registros=len(rows),
    )


def export_audit_log_excel(db: DbSession) -> GeneratedExport:
    rows = _build_audit_log_export_rows(db)
    return GeneratedExport(
        content=generate_excel_bytes(
            (
                ExportSheet(
                    title="Log Auditoria",
                    columns=AUDIT_LOG_EXPORT_COLUMNS,
                    rows=rows,
                ),
            )
        ),
        total_registros=len(rows),
    )
